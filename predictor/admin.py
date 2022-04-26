from django.contrib import admin
from django.contrib.auth.models import Group, User
from .models import *
from django.utils.html import format_html
from django.contrib.auth.admin import UserAdmin
from django.http import HttpResponse
from openpyxl import Workbook
# Register your models here.


@admin.action(description='Export selected to XLSX')
def export(self, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=Heart Datas ('+str(len(queryset))+').xlsx'
    wb=Workbook()
    ws=wb.active
    ws.title='Heart Datas'
    row_num = 1
    columns=['Owner','Age','Gender','Breathelessness during activity','Breathlessness at rest','Breathlessness at night','Exercise induced angina ','Diabetic',
    'Blood Pressure','Cyanosis','Clubbing','Checked Date','Probability']
    print(len(columns))
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
    rows = queryset
    for ro in rows:
        row_num += 1
        row = [ro.owner.name if ro.owner.name else ro.owner.username ,ro.age, ro.get_gender_display(), ro.get_activity_display(), ro.get_rest_display(), ro.get_night_display(), ro.get_exercise_display(), 
        ro.get_diabetes_display(), ro.get_bp_display(), ro.get_cyanosis_display(), ro.get_clubbing_display(), ro.date, ro.probability]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = str(cell_value)
    wb.save(response)
    return response

@admin.action(description='Export selected to XLSX')
def quickexport(self, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=Quick Heart Datas ('+str(len(queryset))+').xlsx'
    wb=Workbook()
    ws=wb.active
    ws.title='Heart Datas'
    row_num = 1
    columns=['Age','Gender','Breathelessness during activity','Breathlessness at rest','Breathlessness at night','Exercise induced angina ','Diabetic',
    'Blood Pressure','Cyanosis','Clubbing','Checked Date','Probability']
    print(len(columns))
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
    rows = queryset
    for ro in rows:
        row_num += 1
        row = [ro.age, ro.get_gender_display(), ro.get_activity_display(), ro.get_rest_display(), ro.get_night_display(), ro.get_exercise_display(), 
        ro.get_diabetes_display(), ro.get_bp_display(), ro.get_cyanosis_display(), ro.get_clubbing_display(), ro.date, ro.probability]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = str(cell_value)
    wb.save(response)
    return response
class HeartDataAdmin(admin.ModelAdmin):
    list_display=['title','age','probability']
    list_filter = ['probability','date','owner']
    actions = [export]
    def has_change_permission(self, request, obj=None):
        return False
    def title(self, obj):
        return format_html(f'<div>{obj.owner.first_name if obj.owner.first_name else obj.owner.username} {obj.owner.last_name if obj.owner.last_name else None}</div>')

admin.site.register(HeartData, HeartDataAdmin)

class QuickHeartDataAdmin(admin.ModelAdmin):
    list_display=['title','age','probability']
    list_filter = ['probability','date']
    actions = [quickexport]
    def has_change_permission(self, request, obj=None):
        return False
    def title(self,obj):
        return f'Quick Heart Data ({obj.id})'

admin.site.register(QuickHeartData, QuickHeartDataAdmin)

admin.site.register(DoctorHospital)
admin.site.unregister(Group)
admin.site.unregister(User)
class NewUserAdmin(UserAdmin):
    #def has_change_permission(self, request, obj=None):
        #return False  
    exclude = ['']

admin.site.register(User, NewUserAdmin)